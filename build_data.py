"""
build_data.py
Mengubah Soal_Latihan_buatan_AI_Persiapan_Responsi_Biokim_Gizi_v2.xlsx
menjadi questions_data.js (window.QUESTIONS = [...]) untuk dipakai index.html.

Cara pakai:
  pip install openpyxl
  python build_data.py

Output: questions_data.js (letakkan satu folder dengan index.html)
"""

import openpyxl
import json
import re

XLSX_PATH = 'Soal_Latihan_buatan_AI_Persiapan_Responsi_Biokim_Gizi_v2.xlsx'
OUTPUT_JS  = 'questions_data.js'

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# ── Sheet 1: Soal Overview ──────────────────────────────────────────────────
questions = []
ws = wb['Soal Overview (40)']
for row in ws.iter_rows(min_row=2, values_only=True):
    no, pertanyaan, jawaban = row[0], row[1], row[2]
    if no is None or pertanyaan is None:
        continue
    questions.append({
        "id":        int(no),
        "group":     "overview",
        "question":  str(pertanyaan).strip(),
        "answer":    str(jawaban).strip() if jawaban else "",
        "needsTable": False,
    })

# ── Sheet 2 & 3: Kasus Hitung / Kasus Lanjutan ─────────────────────────────
def parse_calc_sheet(sheet_name, group_key):
    ws = wb[sheet_name]
    out = []
    for row in ws.iter_rows(values_only=True):
        no = row[0]
        if no is None:
            continue
        try:
            no_int = int(no)
        except (ValueError, TypeError):
            continue
        kasus, pertanyaan, rumus, jawaban = row[1], row[2], row[3], row[4]
        full_q = f"{kasus}\n\n{pertanyaan}".strip()
        out.append({
            "id":        no_int,
            "group":     group_key,
            "question":  full_q,
            "rumus":     str(rumus).strip() if rumus else "",
            "answer":    str(jawaban).strip() if jawaban else "",
            "needsTable": False,
        })
    return out

hitung   = parse_calc_sheet('Soal Kasus Hitung (20)',   'hitung')
lanjutan = parse_calc_sheet('Soal Kasus Lanjutan (20)', 'lanjutan')

# ── Rework: sembunyikan nilai FS pada soal-soal tertentu ──────────────────
# Hanya soal yang menghitung TEE yang menyembunyikan nilai FS (sekitar separuh).
# Deskripsi kondisi/trauma TETAP ditampilkan — hanya angka faktornya yang
# disembunyikan agar mahasiswa harus membaca tabel referensi sendiri.

by_id = {q['id']: q for q in hitung + lanjutan}

REWORK = {
    47: {
        "old": "FS (kanker + kemoterapi) = 1,4.",
        "new": "Faktor stres: kanker + kemoterapi (tentukan nilai faktor stres yang sesuai dari tabel referensi di bawah).",
    },
    55: {
        "old": "FS=1,3.",
        "new": "Faktor stres: tidak ada stres klinis (tentukan nilai faktor stres yang sesuai dari tabel referensi di bawah).",
    },
    63: {
        "old": "FS = 1,3.",
        "new": "Faktor stres: tidak ada kondisi stres klinis (tentukan nilai faktor stres yang sesuai dari tabel referensi di bawah).",
    },
    70: {
        "old": "FS = 1,7 (stres sangat berat: luka bakar dan sepsis).",
        "new": "Faktor stres: stres sangat berat (luka bakar dan sepsis) (tentukan nilai faktor stres yang sesuai dari tabel referensi di bawah).",
    },
    76: {
        "old": "FS = 1,5.",
        "new": "Faktor stres: sepsis, stres sedang (tentukan nilai faktor stres yang sesuai dari tabel referensi di bawah).",
    },
    79: {
        "old": "(stres bedah multisistem, FS=1,6)",
        "new": "(stres bedah multisistem - tentukan nilai faktor stres yang sesuai dari tabel referensi di bawah)",
    },
}

for qid, rule in REWORK.items():
    if qid not in by_id:
        print(f"[WARNING] ID {qid} tidak ditemukan, dilewati.")
        continue
    q = by_id[qid]
    if rule["old"] not in q["question"]:
        print(f"[WARNING] Pola tidak ditemukan di soal {qid}: {rule['old']!r}")
        continue
    q["question"]  = q["question"].replace(rule["old"], rule["new"])
    q["needsTable"] = True

# ── Manual override jawaban ─────────────────────────────────────────────────
all_q = questions + hitung + lanjutan
by_id_all = {q["id"]: q for q in all_q}

ANSWER_OVERRIDES = {
    52: "Kebutuhan air = √[(160 × 54) / 3600] × 1500 = √(8640 / 3600) × 1500 = √2.4 × 1500 ≈ 1.549 × 1500 ≈ 2.324 mL/hari",
    60: "Kebutuhan cairan = 35 mL × 54 kg = 1.890 mL/hari",
    61: "IMT = 70 / (1,70)² = 24,22 kg/m² → Kategori: Overweight (Kelebihan berat badan) berdasarkan klasifikasi Asia-Pasifik (≥ 23 kg/m²).",
    80: "Kebutuhan cairan = 35 mL × 52 kg = 1.820 mL/hari",
}

for qid, ans in ANSWER_OVERRIDES.items():
    if qid in by_id_all:
        by_id_all[qid]["answer"] = ans

# ── Tulis output ────────────────────────────────────────────────────────────
with open(OUTPUT_JS, 'w', encoding='utf-8') as f:
    f.write("window.QUESTIONS = ")
    json.dump(all_q, f, ensure_ascii=False, indent=2)
    f.write(";")

ov  = len(questions)
hit = len(hitung)
lan = len(lanjutan)
print(f"✓ {OUTPUT_JS} ditulis. Total: {ov+hit+lan} soal (Overview: {ov}, Hitung: {hit}, Lanjutan: {lan})")
